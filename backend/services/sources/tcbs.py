import datetime
import io
import re
from typing import Dict, List, Optional

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

try:
    import xlrd
except ImportError:
    xlrd = None

from models import Asset
from services.sources.base import Source
from services.sources.utils import parse_float, today


TCBS_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9,vi-VN;q=0.8,vi;q=0.7",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
}

TCBS_URL = "https://www.tcbs.com.vn"
TCC_URL = "https://www.techcomcapital.com.vn"

# Fund symbol -> TCBS iFund slug
FUND_SLUGS = {
    "TCBF": "tcbf",
    "TCEF": "tcef",
    "TCFF": "tcff",
    "TCFIN": "tcfin",
    "TCRES": "tcres",
    "TCSME": "tcsme",
    "FUETCC50": "quy-etf-techcom-capital-vnx50",
    "TCREIT": "tcreit",
}

# Fund symbol -> Techcom Capital detail page slug
FUND_DETAIL_SLUGS = {
    "TCBF": "quy-dau-tu-trai-phieu-tcbf",
    "TCEF": "quy-dau-tu-co-phieu-techcom-tcef",
    "TCFF": "quy-dau-tu-trai-phieu-linh-hoat-techcom-tcff",
    "TCFIN": "quy-dau-tu-co-phieu-ngan-hang-va-tai-chinh-techcom-tcfin",
    "TCRES": "quy-dau-tu-co-phieu-bat-dong-san-techcom-tcres",
    "TCSME": "quy-dau-tu-co-phieu-doanh-nghiep-vua-va-nho-techcom-tcsme",
    "FUETCC50": "quy-dau-tu-etf-techcom-capital-vnx50-fuetcc50",
}

# Fund symbol -> Techcom Capital disclosure page slug
FUND_DISCLOSURE_SLUGS = {
    "TCBF": "tai-lieu-quy-tcbf",
    "TCEF": "tai-lieu-quy-tcef",
    "TCFF": "tai-lieu-quy-tcff",
    "TCFIN": "tai-lieu-quy-tcfin",
    "TCRES": "tai-lieu-quy-tcres",
    "TCSME": "tai-lieu-quy-tcsme",
    "FUETCC50": "tai-lieu-quy-fuetcc50",
    "TCREIT": "cong-bo-thong-tin-quy-tcreit",
}


class TcbsSource(Source):
    """Dữ liệu cổ phiếu/quỹ/ETF từ TCBS (Techcom Securities).

    Stock data: Uses the TCBS public API (apipubaws.tcbs.com.vn) for price
    history. The API may be unavailable; the source will fall back to other
    sources in the registry via SourceSelector.

    Fund data: Scrapes the TCBS iFund page (tcbs.com.vn) for fund listings
    and the Techcom Capital website (techcomcapital.com.vn) for NAV reports.
    NAV values are extracted from Excel files linked on daily NAV report
    posts on the disclosure pages.
    """

    code = "tcbs"
    name = "TCBS"
    description = "Dữ liệu cổ phiếu/quỹ/ETF từ TCBS (Techcom Securities)."
    supported_types = ["STOCK", "FUND", "ETF"]
    supports_history = True
    supports_listing = True

    # --- Caching ---
    _FUND_LISTING_CACHE: Optional[List[dict]] = None
    _FUND_LISTING_CACHE_TIME: Optional[datetime.datetime] = None
    _NAV_REPORT_CACHE: Dict[str, List[dict]] = {}
    _NAV_REPORT_CACHE_TIME: Dict[str, datetime.datetime] = {}

    # ------------------------------------------------------------------
    # Stock: API-based price history (legacy, may be unavailable)
    # ------------------------------------------------------------------
    def _fetch_bars(
        self, symbol: str, start: datetime.date, end: datetime.date
    ) -> Dict[datetime.date, float]:
        try:
            end_dt = datetime.datetime.combine(end, datetime.time.max)
            to_ts = int(end_dt.timestamp())
            days = (end - start).days + 1
            url = (
                "https://apipubaws.tcbs.com.vn/stock-insight/v2/stock/bars-long-term"
                f"?ticker={symbol.upper()}&type=stock&resolution=D"
                f"&to={to_ts}&countBack={days}"
            )
            r = requests.get(url, headers=TCBS_HEADERS, timeout=15)
            if r.status_code == 200:
                data = r.json().get("data", [])
                result = {}
                for row in data:
                    trading_date = row.get("tradingDate")
                    close = row.get("close")
                    if not trading_date or close is None:
                        continue
                    d = _parse_trading_date(trading_date)
                    if d and start <= d <= end:
                        result[d] = parse_float(close)
                if result:
                    return result
        except Exception as e:
            print(f"[source tcbs] bars {symbol} error: {e}")
        return {}

    # ------------------------------------------------------------------
    # Fund: HTML scraping from TCBS + Techcom Capital
    # ------------------------------------------------------------------
    def _fetch_fund_listing(self, force: bool = False) -> List[dict]:
        """Scrape fund listing from TCBS iFund page."""
        if (
            not force
            and self._FUND_LISTING_CACHE is not None
            and self._FUND_LISTING_CACHE_TIME is not None
            and (datetime.datetime.now() - self._FUND_LISTING_CACHE_TIME).total_seconds() < 3600
        ):
            return self._FUND_LISTING_CACHE
        try:
            r = requests.get(f"{TCBS_URL}/ca-nhan/ifund/", headers=TCBS_HEADERS, timeout=15)
            if r.status_code != 200:
                return self._FUND_LISTING_CACHE or []
            soup = BeautifulSoup(r.text, "html.parser")
            funds = []
            seen = set()
            for a in soup.find_all("a", href=True):
                href = a["href"]
                text = a.get_text(strip=True)
                if "/ca-nhan/ifund/" not in href or href.endswith("ifund/") or href == "/ca-nhan/ifund/":
                    continue
                if href in seen or not text or text == "Tìm hiểu thêm":
                    continue
                seen.add(href)
                symbol_match = re.search(r"\(([A-Z]{3,12})\)", text)
                symbol = symbol_match.group(1) if symbol_match else ""
                if not symbol:
                    # Try to extract from URL slug
                    slug = href.rstrip("/").split("/")[-1]
                    for sym, sl in FUND_SLUGS.items():
                        if sl == slug:
                            symbol = sym
                            break
                if not symbol:
                    continue
                full_url = href if href.startswith("http") else TCBS_URL + href
                funds.append(
                    {
                        "symbol": symbol,
                        "name": text,
                        "url": full_url,
                    }
                )
            if funds:
                TcbsSource._FUND_LISTING_CACHE = funds
                TcbsSource._FUND_LISTING_CACHE_TIME = datetime.datetime.now()
            return funds or self._FUND_LISTING_CACHE or []
        except Exception as e:
            print(f"[source tcbs] fund listing error: {e}")
            return self._FUND_LISTING_CACHE or []

    def _fetch_nav_report_links(self, symbol: str, force: bool = False) -> List[dict]:
        """Scrape NAV report links from Techcom Capital disclosure page.

        Returns list of {date, title, url} sorted by date descending.
        """
        if (
            not force
            and symbol in self._NAV_REPORT_CACHE
            and symbol in self._NAV_REPORT_CACHE_TIME
            and (datetime.datetime.now() - self._NAV_REPORT_CACHE_TIME[symbol]).total_seconds() < 3600
        ):
            return self._NAV_REPORT_CACHE[symbol]
        slug = FUND_DISCLOSURE_SLUGS.get(symbol.upper())
        if not slug:
            return []
        try:
            url = f"{TCC_URL}/{slug}/"
            r = requests.get(url, headers=TCBS_HEADERS, timeout=15)
            if r.status_code != 200:
                return self._NAV_REPORT_CACHE.get(symbol, [])
            soup = BeautifulSoup(r.text, "html.parser")
            reports = []
            for h in soup.find_all(["h4", "h3", "h2"]):
                text = h.get_text(strip=True)
                if "Giá trị tài sản ròng" not in text and "Net Asset Value" not in text:
                    continue
                a = h.find("a", href=True)
                if not a:
                    continue
                date_match = re.search(r"(\d{2})/(\d{2})/(\d{4})", text)
                date_str = ""
                report_date = None
                if date_match:
                    try:
                        report_date = datetime.date(
                            int(date_match.group(3)),
                            int(date_match.group(2)),
                            int(date_match.group(1)),
                        )
                        date_str = date_match.group(0)
                    except ValueError:
                        pass
                reports.append(
                    {
                        "date": report_date,
                        "date_str": date_str,
                        "title": text,
                        "url": a["href"],
                    }
                )
            reports.sort(key=lambda x: x["date"] or datetime.date.min, reverse=True)
            if reports:
                TcbsSource._NAV_REPORT_CACHE[symbol] = reports
                TcbsSource._NAV_REPORT_CACHE_TIME[symbol] = datetime.datetime.now()
            return reports or self._NAV_REPORT_CACHE.get(symbol, [])
        except Exception as e:
            print(f"[source tcbs] nav links {symbol} error: {e}")
            return self._NAV_REPORT_CACHE.get(symbol, [])

    def _fetch_excel_nav(self, report_url: str) -> Optional[dict]:
        """Follow a NAV report post URL, find Excel links, download and parse.

        Returns {nav_per_ccq, total_nav, date, num_ccq} or None.
        Handles both .xlsx (openpyxl) and .xls (xlrd) formats.
        """
        try:
            r = requests.get(report_url, headers=TCBS_HEADERS, timeout=15)
            if r.status_code != 200:
                return None
            soup = BeautifulSoup(r.text, "html.parser")
            article = soup.find("article")
            if not article:
                return None
            # Find Excel file links — prefer daily ("ngay") over weekly ("tuan")
            xlsx_url = None
            xlsx_urls = []
            for a in article.find_all("a", href=True):
                href = a["href"]
                if href.endswith(".xlsx") or href.endswith(".xls"):
                    xlsx_urls.append(href)
                    link_text = a.get_text(strip=True).lower()
                    if "ngay" in href.lower() or "ngay" in link_text:
                        xlsx_url = href
                        break
            if not xlsx_url and xlsx_urls:
                xlsx_url = xlsx_urls[0]
            if not xlsx_url:
                return None
            # Download Excel file
            r2 = requests.get(xlsx_url, headers=TCBS_HEADERS, timeout=30)
            if r2.status_code != 200:
                return None
            content = r2.content
            # Parse based on file type
            if xlsx_url.endswith(".xls") and not xlsx_url.endswith(".xlsx"):
                # Old .xls format — use xlrd
                if xlrd is None:
                    print("[source tcbs] xlrd not installed, cannot parse .xls files")
                    return None
                wb = xlrd.open_workbook(file_contents=content)
                return _parse_nav_xls(wb)
            else:
                # .xlsx format — use openpyxl
                wb = load_workbook(io.BytesIO(content), data_only=True)
                return _parse_nav_excel(wb)
        except Exception as e:
            print(f"[source tcbs] excel nav error: {e}")
            return None

    def _fetch_fund_price(self, symbol: str) -> Optional[dict]:
        """Get latest fund NAV price by scraping NAV reports."""
        reports = self._fetch_nav_report_links(symbol)
        if not reports:
            return None
        # Try the latest few reports
        for report in reports[:3]:
            nav_data = self._fetch_excel_nav(report["url"])
            if nav_data and nav_data.get("nav_per_ccq"):
                price = nav_data["nav_per_ccq"]
                prev_price = nav_data.get("prev_nav_per_ccq") or price
                change = price - prev_price
                change_percent = (change / prev_price * 100) if prev_price else 0.0
                return {
                    "price": price,
                    "change": change,
                    "change_percent": change_percent,
                    "date": nav_data.get("date") or report["date"],
                    "metadata": {
                        "source": "tcbs",
                        "total_nav": nav_data.get("total_nav"),
                        "num_ccq": nav_data.get("num_ccq"),
                    },
                }
        return None

    def _fetch_fund_history(
        self, symbol: str, start: datetime.date, end: datetime.date
    ) -> Dict[datetime.date, float]:
        """Get fund NAV history by scraping multiple NAV reports."""
        reports = self._fetch_nav_report_links(symbol)
        if not reports:
            return {}
        result = {}
        for report in reports:
            report_date = report.get("date")
            if not report_date or report_date < start or report_date > end:
                continue
            nav_data = self._fetch_excel_nav(report["url"])
            if nav_data and nav_data.get("nav_per_ccq"):
                nav_date = nav_data.get("date") or report_date
                if start <= nav_date <= end:
                    result[nav_date] = nav_data["nav_per_ccq"]
        return result

    # ------------------------------------------------------------------
    # Source interface methods
    # ------------------------------------------------------------------
    def fetch_price(self, asset: Asset) -> Optional[dict]:
        if asset.type == "FUND":
            return self._fetch_fund_price(asset.symbol)
        # Stock/ETF: use API-based bars
        try:
            end = today()
            start = end - datetime.timedelta(days=14)
            history = self._fetch_bars(asset.symbol, start, end)
            if not history:
                return None
            latest_date = max(history.keys())
            latest = history[latest_date]
            prev_date = sorted(history.keys())[-2] if len(history) > 1 else latest_date
            prev = history[prev_date]
            change = latest - prev
            change_percent = (change / prev * 100) if prev else 0.0
            return {
                "price": latest,
                "change": change,
                "change_percent": change_percent,
                "date": latest_date,
                "metadata": {"source": "tcbs"},
            }
        except Exception as e:
            print(f"[source tcbs] price {asset.symbol} error: {e}")
        return None

    def fetch_history(
        self,
        symbol: str,
        asset_type: str,
        start: datetime.date,
        end: datetime.date,
    ) -> Dict[datetime.date, float]:
        if asset_type == "FUND":
            return self._fetch_fund_history(symbol, start, end)
        return self._fetch_bars(symbol, start, end)

    def fetch_listing(self) -> List[dict]:
        results = []
        # --- Funds: scrape from TCBS iFund page ---
        try:
            funds = self._fetch_fund_listing()
            for fund in funds:
                results.append(
                    {
                        "symbol": fund["symbol"],
                        "name": fund["name"],
                        "exchange": "TCBS",
                        "type": "FUND",
                        "metadata": {"url": fund.get("url")},
                    }
                )
        except Exception as e:
            print(f"[source tcbs] fund listing error: {e}")
        # --- Stocks: use API (may be unavailable) ---
        try:
            offset = 0
            limit = 1000
            while True:
                url = (
                    "https://apipubaws.tcbs.com.vn/tcanalysis/v1/ticker/all"
                    f"?offset={offset}&limit={limit}"
                )
                r = requests.get(url, headers=TCBS_HEADERS, timeout=15)
                if r.status_code != 200:
                    break
                items = r.json().get("data", [])
                if not items:
                    break
                for item in items:
                    symbol = item.get("ticker")
                    name = item.get("org_name") or item.get("short_name") or symbol
                    exchange = item.get("exchange") or ""
                    if not symbol:
                        continue
                    results.append(
                        {
                            "symbol": symbol,
                            "name": name,
                            "exchange": exchange,
                            "type": "STOCK",
                            "metadata": item,
                        }
                    )
                if len(items) < limit:
                    break
                offset += limit
        except Exception as e:
            print(f"[source tcbs] stock listing error: {e}")
        return results


# ------------------------------------------------------------------
# Helpers
# ------------------------------------------------------------------
def _parse_trading_date(value) -> Optional[datetime.date]:
    if isinstance(value, datetime.date):
        return value
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, (int, float)):
        try:
            return datetime.datetime.fromtimestamp(value / 1000).date()
        except Exception:
            pass
    for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.datetime.strptime(str(value), fmt).date()
        except Exception:
            continue
    return None


def _parse_nav_excel(wb) -> Optional[dict]:
    """Parse a Techcom Capital NAV Excel workbook.

    Expected structure (PL 24 Daily sheet):
    - Row 5: Date line "Tại ngày DD tháng MM năm YYYY / As at DD Mon YYYY"
    - Row 18: Total NAV (col D = this period, col E = last period)
    - Row 20: NAV per Fund Certificate (col D = this period, col E = last period)
    - Row 22: Number of Fund Certificates (col D = this period)
    """
    try:
        ws = wb[wb.sheetnames[0]]
        nav_date = None
        nav_per_ccq = None
        prev_nav_per_ccq = None
        total_nav = None
        num_ccq = None

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            if not any(cell is not None for cell in row):
                continue
            row_text = " ".join(str(c) for c in row if c is not None)

            # Parse date from row 5: "Tại ngày 06 tháng 08 năm 2026"
            if "Tại ngày" in row_text or "As at" in row_text:
                date_match = re.search(r"(\d{1,2})\s*tháng\s*(\d{1,2})\s*năm\s*(\d{4})", row_text)
                if date_match:
                    try:
                        nav_date = datetime.date(
                            int(date_match.group(3)),
                            int(date_match.group(2)),
                            int(date_match.group(1)),
                        )
                    except ValueError:
                        pass
                if not nav_date:
                    # Try English format: "As at 06 Aug 2026"
                    en_match = re.search(r"As at\s*(\d{1,2})\s+(\w{3})\s+(\d{4})", row_text)
                    if en_match:
                        try:
                            nav_date = datetime.datetime.strptime(
                                f"{en_match.group(1)} {en_match.group(2)} {en_match.group(3)}",
                                "%d %b %Y",
                            ).date()
                        except ValueError:
                            pass

            # Row with "của một chứng chỉ quỹ/ per Fund Certificate"
            if "per Fund Certificate" in row_text or "của một chứng chỉ quỹ" in row_text:
                label_idx = -1
                for i, c in enumerate(row):
                    if c is not None and ("per Fund Certificate" in str(c) or "của một chứng chỉ quỹ" in str(c)):
                        label_idx = i
                        break
                nums = []
                if label_idx >= 0:
                    for c in row[label_idx + 1:]:
                        if c is None or (isinstance(c, str) and not c.strip()):
                            continue
                        v = parse_float(c)
                        if v is not None and v > 0:
                            nums.append(v)
                if nums:
                    nav_per_ccq = nums[0]
                if len(nums) > 1:
                    prev_nav_per_ccq = nums[1]

            # Row with "của quỹ/ of the Fund" (total NAV)
            if ("of the Fund" in row_text or "của quỹ" in row_text) and "per" not in row_text.lower():
                label_idx = -1
                for i, c in enumerate(row):
                    if c is not None and ("of the Fund" in str(c) or "của quỹ" in str(c)):
                        label_idx = i
                        break
                nums = []
                if label_idx >= 0:
                    for c in row[label_idx + 1:]:
                        if c is None or (isinstance(c, str) and not c.strip()):
                            continue
                        v = parse_float(c)
                        if v is not None and v > 0:
                            nums.append(v)
                if nums:
                    total_nav = nums[0]

            # Row with "Number of Fund Certificates"
            if "Number of Fund Certificates" in row_text or "Số lượng chứng chỉ quỹ" in row_text:
                label_idx = -1
                for i, c in enumerate(row):
                    if c is not None and ("Number of Fund Certificates" in str(c) or "Số lượng chứng chỉ quỹ" in str(c)):
                        label_idx = i
                        break
                nums = []
                if label_idx >= 0:
                    for c in row[label_idx + 1:]:
                        if c is None or (isinstance(c, str) and not c.strip()):
                            continue
                        v = parse_float(c)
                        if v is not None and v > 0:
                            nums.append(v)
                if nums:
                    num_ccq = nums[0]

        if nav_per_ccq and nav_per_ccq > 0:
            return {
                "nav_per_ccq": nav_per_ccq,
                "prev_nav_per_ccq": prev_nav_per_ccq,
                "total_nav": total_nav,
                "num_ccq": num_ccq,
                "date": nav_date,
            }
    except Exception as e:
        print(f"[source tcbs] parse nav excel error: {e}")
    return None


def _parse_nav_xls(wb) -> Optional[dict]:
    """Parse a Techcom Capital NAV Excel workbook in old .xls format (xlrd).

    Same structure as _parse_nav_excel but uses xlrd's API.
    """
    try:
        ws = wb.sheet_by_index(0)
        nav_date = None
        nav_per_ccq = None
        prev_nav_per_ccq = None
        total_nav = None
        num_ccq = None

        for row_idx in range(ws.nrows):
            row = [ws.cell_value(row_idx, col) for col in range(ws.ncols)]
            if not any(c not in (None, "", 0) for c in row if not isinstance(c, str)):
                row_text = " ".join(str(c) for c in row if c != "")
            else:
                row_text = " ".join(str(c) for c in row if c is not None and str(c).strip() != "")
            if not row_text.strip():
                continue

            # Parse date
            if "Tại ngày" in row_text or "As at" in row_text or "Từ ngày" in row_text:
                date_match = re.search(r"(\d{1,2})\s*tháng\s*(\d{1,2})\s*năm\s*(\d{4})", row_text)
                if date_match:
                    try:
                        nav_date = datetime.date(
                            int(date_match.group(3)),
                            int(date_match.group(2)),
                            int(date_match.group(1)),
                        )
                    except ValueError:
                        pass
                if not nav_date:
                    en_match = re.search(r"As at\s*(\d{1,2})\s+(\w{3})\s+(\d{4})", row_text)
                    if en_match:
                        try:
                            nav_date = datetime.datetime.strptime(
                                f"{en_match.group(1)} {en_match.group(2)} {en_match.group(3)}",
                                "%d %b %Y",
                            ).date()
                        except ValueError:
                            pass
                if not nav_date:
                    tu_match = re.search(r"Từ ngày\s*(\d{1,2})/(\d{1,2})/(\d{4})", row_text)
                    if tu_match:
                        try:
                            nav_date = datetime.date(
                                int(tu_match.group(3)),
                                int(tu_match.group(2)),
                                int(tu_match.group(1)),
                            )
                        except ValueError:
                            pass
                if not nav_date:
                    for c in row:
                        if isinstance(c, (int, float)) and not isinstance(c, bool) and c > 255:
                            try:
                                nav_date = datetime.date.fromordinal(
                                    datetime.date(1899, 12, 30).toordinal() + int(c)
                                )
                                break
                            except ValueError:
                                continue

            # NAV per Fund Certificate
            if "per Fund Certificate" in row_text or "của một chứng chỉ quỹ" in row_text:
                label_idx = -1
                for i, c in enumerate(row):
                    if c not in (None, "") and ("per Fund Certificate" in str(c) or "của một chứng chỉ quỹ" in str(c)):
                        label_idx = i
                        break
                nums = []
                if label_idx >= 0:
                    for c in row[label_idx + 1:]:
                        if c in (None, "") or (isinstance(c, str) and not c.strip()):
                            continue
                        v = parse_float(c)
                        if v is not None and v > 0:
                            nums.append(v)
                if nums:
                    nav_per_ccq = nums[0]
                if len(nums) > 1:
                    prev_nav_per_ccq = nums[1]

            # Total NAV
            if ("of the Fund" in row_text or "của quỹ" in row_text) and "per" not in row_text.lower():
                label_idx = -1
                for i, c in enumerate(row):
                    if c not in (None, "") and ("of the Fund" in str(c) or "của quỹ" in str(c)):
                        label_idx = i
                        break
                nums = []
                if label_idx >= 0:
                    for c in row[label_idx + 1:]:
                        if c in (None, "") or (isinstance(c, str) and not c.strip()):
                            continue
                        v = parse_float(c)
                        if v is not None and v > 0:
                            nums.append(v)
                if nums:
                    total_nav = nums[0]

            # Number of Fund Certificates
            if "Number of Fund Certificates" in row_text or "Số lượng chứng chỉ quỹ" in row_text:
                label_idx = -1
                for i, c in enumerate(row):
                    if c not in (None, "") and ("Number of Fund Certificates" in str(c) or "Số lượng chứng chỉ quỹ" in str(c)):
                        label_idx = i
                        break
                nums = []
                if label_idx >= 0:
                    for c in row[label_idx + 1:]:
                        if c in (None, "") or (isinstance(c, str) and not c.strip()):
                            continue
                        v = parse_float(c)
                        if v is not None and v > 0:
                            nums.append(v)
                if nums:
                    num_ccq = nums[0]

        if nav_per_ccq and nav_per_ccq > 0:
            return {
                "nav_per_ccq": nav_per_ccq,
                "prev_nav_per_ccq": prev_nav_per_ccq,
                "total_nav": total_nav,
                "num_ccq": num_ccq,
                "date": nav_date,
            }
    except Exception as e:
        print(f"[source tcbs] parse nav xls error: {e}")
    return None
