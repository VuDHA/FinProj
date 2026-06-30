import csv
import io
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from sqlmodel import Session

from config import settings
from models import Asset, Transaction
from schemas import CsvImportResult
from services.csv_io import VALID_ASSET_TYPES
from services.ollama_client import OllamaClient, OllamaClientError


ASSET_TARGET_FIELDS = ["symbol", "name", "type", "exchange", "currency"]
REQUIRED_ASSET_FIELDS = ["symbol", "name", "type"]
TRANSACTION_TARGET_FIELDS = ["symbol", "type", "quantity", "price", "fee", "date", "notes"]
REQUIRED_TRANSACTION_FIELDS = ["symbol", "type", "quantity", "price", "date"]


class SmartImportService:
    """Preview and import CSV/Excel files with AI-assisted header mapping."""

    def __init__(self, model: str = settings.OLLAMA_MODEL):
        self.model = model
        self._client = OllamaClient()

    @staticmethod
    def _is_csv(filename: str) -> bool:
        return filename.lower().endswith(".csv")

    @staticmethod
    def _is_excel(filename: str) -> bool:
        return filename.lower().endswith((".xlsx", ".xls"))

    def _read_csv(self, content: bytes) -> List[Dict[str, Any]]:
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        return [dict(row) for row in reader]

    def _read_excel(
        self, content: bytes, sheet_name: Optional[str] = None
    ) -> List[Dict[str, Any]]:
        wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
        if sheet_name is None:
            sheet = wb.active
        else:
            sheet = wb[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(c or "").strip() for c in rows[0]]
        return [
            {
                headers[i]: (str(c) if c is not None else "")
                for i, c in enumerate(row)
                if i < len(headers) and headers[i]
            }
            for row in rows[1:]
        ]

    def _read_rows(
        self, content: bytes, filename: str, sheet_name: Optional[str] = None
    ) -> List[Dict[str, Any]]:
        if self._is_csv(filename):
            return self._read_csv(content)
        if self._is_excel(filename):
            return self._read_excel(content, sheet_name=sheet_name)
        raise ValueError("Unsupported file format. Only .csv and .xlsx are supported.")

    def preview(
        self, content: bytes, filename: str, sheet_name: Optional[str] = None
    ) -> Dict[str, Any]:
        """Return sheet names, headers, and sample rows for a file."""
        rows = self._read_rows(content, filename, sheet_name=sheet_name)
        headers = list(rows[0].keys()) if rows else []
        sample = rows[:5]

        sheet_names = None
        actual_sheet = sheet_name
        if self._is_excel(filename):
            wb = load_workbook(filename=io.BytesIO(content), read_only=True)
            sheet_names = wb.sheetnames
            if actual_sheet is None:
                actual_sheet = wb.active.title if wb.active else None

        return {
            "filename": filename,
            "sheet_names": sheet_names,
            "sheet": actual_sheet,
            "headers": headers,
            "sample_rows": sample,
            "row_count": len(rows),
        }

    def _build_mapping_prompt(
        self, headers: List[str], import_type: str, language: str = "vi"
    ) -> str:
        target_fields = (
            ASSET_TARGET_FIELDS if import_type == "assets" else TRANSACTION_TARGET_FIELDS
        )
        if language == "vi":
            return (
                "Bạn là trợ lý tài chính. Bạn nhận được danh sách tiêu đề cột từ file CSV/Excel. "
                "Hãy ánh xạ mỗi tiêu đề sang một trường đích phù hợp. "
                "Trả về một JSON object duy nhất với key là tiêu đề gốc và value là trường đích. "
                "Nếu không khớp, dùng null. Không thêm giải thích.\n\n"
                f"Loại import: {import_type}\n"
                f"Trường đích: {', '.join(target_fields)}\n"
                f"Tiêu đề: {', '.join(headers)}\n\n"
                "JSON:"
            )
        return (
            "You are a financial assistant. Map each source column header to a target field. "
            "Return a single JSON object where keys are the original headers and values are target fields. "
            "Use null if no match. No commentary.\n\n"
            f"Import type: {import_type}\n"
            f"Target fields: {', '.join(target_fields)}\n"
            f"Headers: {', '.join(headers)}\n\n"
            "JSON:"
        )

    def suggest_mapping(
        self, headers: List[str], import_type: str, language: str = "vi"
    ) -> Dict[str, Optional[str]]:
        """Use the local LLM to suggest a header-to-target mapping."""
        if not settings.OLLAMA_ENABLED:
            return {h: self._fallback_mapping(h, import_type) for h in headers}

        target_fields = (
            ASSET_TARGET_FIELDS if import_type == "assets" else TRANSACTION_TARGET_FIELDS
        )
        try:
            raw = self._client.generate(
                prompt=self._build_mapping_prompt(headers, import_type, language),
                model=self.model,
                options={
                    "temperature": 0.1,
                    "num_predict": 256,
                },
                task_name="smart_import_mapping",
            )
        except OllamaClientError as e:
            print(f"[smart_import] mapping failed: {e}")
            return {h: self._fallback_mapping(h, import_type) for h in headers}

        try:
            import json

            mapping = json.loads(raw.strip())
        except json.JSONDecodeError as e:
            print(f"[smart_import] mapping JSON parse failed: {e}")
            return {h: self._fallback_mapping(h, import_type) for h in headers}

        # Validate values against target fields.
        cleaned = {}
        for header in headers:
            value = mapping.get(header)
            if value not in target_fields:
                value = None
            cleaned[header] = value
        return cleaned

    @staticmethod
    def _fallback_mapping(header: str, import_type: str) -> Optional[str]:
        """Simple keyword-based fallback mapping."""
        h = header.lower().strip()
        target_fields = (
            ASSET_TARGET_FIELDS if import_type == "assets" else TRANSACTION_TARGET_FIELDS
        )
        for field in target_fields:
            if field in h or h in field:
                return field
            # Vietnamese common aliases
            aliases = {
                "symbol": ["mã", "ticker", "mã cp"],
                "name": ["tên", "tên tài sản", "tên cp"],
                "type": ["loại", "loại tài sản"],
                "exchange": ["sàn"],
                "currency": ["tiền tệ", "đơn vị tiền"],
                "quantity": ["số lượng", "khối lượng"],
                "price": ["giá", "đơn giá"],
                "fee": ["phí", "hoa hồng"],
                "date": ["ngày"],
                "notes": ["ghi chú", "mô tả"],
            }
            if field in aliases and any(alias in h for alias in aliases[field]):
                return field
        return None

    def import_data(
        self,
        session: Session,
        content: bytes,
        filename: str,
        import_type: str,
        mapping: Dict[str, Optional[str]],
        sheet_name: Optional[str] = None,
    ) -> CsvImportResult:
        """Import a file using the provided header-to-target mapping."""
        rows = self._read_rows(content, filename, sheet_name=sheet_name)
        if not rows:
            return CsvImportResult(created=0, skipped=0, errors=["No rows found in file"])

        target_fields = (
            ASSET_TARGET_FIELDS if import_type == "assets" else TRANSACTION_TARGET_FIELDS
        )
        required_fields = (
            REQUIRED_ASSET_FIELDS if import_type == "assets" else REQUIRED_TRANSACTION_FIELDS
        )
        # Invert mapping: target -> source header
        target_to_source = {v: k for k, v in mapping.items() if v in target_fields}
        missing = [f for f in required_fields if f not in target_to_source]
        if missing:
            return CsvImportResult(
                created=0,
                skipped=0,
                errors=[f"Missing required mappings: {', '.join(missing)}"],
            )

        if import_type == "assets":
            return self._import_assets(session, rows, target_to_source)
        if import_type == "transactions":
            return self._import_transactions(session, rows, target_to_source)
        return CsvImportResult(created=0, skipped=0, errors=["Invalid import type"])

    def _import_assets(
        self, session: Session, rows: List[Dict[str, Any]], target_to_source: Dict[str, str]
    ) -> CsvImportResult:
        created = 0
        skipped = 0
        errors: List[str] = []
        for i, row in enumerate(rows, start=2):
            symbol = (row.get(target_to_source.get("symbol", "")) or "").strip().upper()
            name = (row.get(target_to_source.get("name", "")) or "").strip()
            type_ = (row.get(target_to_source.get("type", "")) or "").strip().upper()
            if not symbol or not name or not type_:
                errors.append(f"Row {i}: missing symbol/name/type")
                skipped += 1
                continue
            if type_ not in VALID_ASSET_TYPES:
                errors.append(f"Row {i}: invalid asset type {type_}")
                skipped += 1
                continue
            existing = session.exec(
                select(Asset).where(Asset.symbol == symbol, Asset.is_active == True)
            ).first()
            if existing:
                skipped += 1
                continue
            asset = Asset(
                symbol=symbol,
                name=name,
                type=type_,
                exchange=(row.get(target_to_source.get("exchange", "")) or "").strip() or None,
                currency=(row.get(target_to_source.get("currency", "")) or "VND").strip() or "VND",
            )
            session.add(asset)
            created += 1
        session.commit()
        return CsvImportResult(created=created, skipped=skipped, errors=errors)

    def _import_transactions(
        self, session: Session, rows: List[Dict[str, Any]], target_to_source: Dict[str, str]
    ) -> CsvImportResult:
        created = 0
        skipped = 0
        errors: List[str] = []
        import datetime

        for i, row in enumerate(rows, start=2):
            symbol = (row.get(target_to_source.get("symbol", "")) or "").strip().upper()
            type_ = (row.get(target_to_source.get("type", "")) or "").strip().upper()
            if not symbol or type_ not in ("BUY", "SELL"):
                errors.append(f"Row {i}: invalid symbol or type")
                skipped += 1
                continue
            try:
                quantity = float(
                    (row.get(target_to_source.get("quantity", "")) or "0")
                    .replace(",", "")
                )
                price = float(
                    (row.get(target_to_source.get("price", "")) or "0")
                    .replace(",", "")
                )
                fee = float(
                    (row.get(target_to_source.get("fee", "")) or "0")
                    .replace(",", "")
                )
                date = datetime.date.fromisoformat(
                    (row.get(target_to_source.get("date", "")) or "").strip()
                )
            except Exception as e:
                errors.append(f"Row {i}: invalid numeric/date field ({e})")
                skipped += 1
                continue

            if quantity <= 0 or price < 0 or fee < 0:
                errors.append(f"Row {i}: quantity/price/fee must be positive")
                skipped += 1
                continue

            asset = session.exec(
                select(Asset).where(Asset.symbol == symbol, Asset.is_active == True)
            ).first()
            if not asset:
                asset = Asset(symbol=symbol, name=symbol, type="STOCK")
                session.add(asset)
                session.commit()
                session.refresh(asset)

            if type_ == "SELL":
                existing = session.exec(
                    select(Transaction).where(Transaction.asset_id == asset.id)
                ).all()
                holding = sum(
                    t.quantity if t.type == "BUY" else -t.quantity for t in existing
                )
                if quantity > holding:
                    errors.append(f"Row {i}: cannot sell {quantity}, holding is {holding}")
                    skipped += 1
                    continue

            tx = Transaction(
                asset_id=asset.id,
                type=type_,
                quantity=quantity,
                price=price,
                fee=fee,
                date=date,
                notes=(row.get(target_to_source.get("notes", "")) or "").strip() or None,
            )
            session.add(tx)
            created += 1
        session.commit()
        return CsvImportResult(created=created, skipped=skipped, errors=errors)
