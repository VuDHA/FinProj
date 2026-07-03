import csv
import io
import zipfile
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook


def is_csv(filename: str) -> bool:
    return filename.lower().endswith(".csv")


def is_excel(filename: str) -> bool:
    return filename.lower().endswith(".xlsx")


def is_zip(filename: str) -> bool:
    return filename.lower().endswith(".zip")


def is_supported_import(filename: str) -> bool:
    return is_csv(filename) or is_excel(filename) or is_zip(filename)


def read_csv(content: bytes) -> List[Dict[str, Any]]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]


def read_excel(content: bytes, sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
    try:
        wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"Cannot read Excel file: {e}") from e
    if sheet_name is None:
        sheet = wb.active
    else:
        sheet = wb[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(c or "").strip() for c in rows[0]]
    return [
        {
            headers[i]: (str(c) if c is not None else "")
            for i, c in enumerate(row)
            if i < len(headers) and headers[i]
        }
        for row in rows[1:]
    ]


def read_excel_sheet_names(content: bytes) -> List[str]:
    try:
        wb = load_workbook(filename=io.BytesIO(content), read_only=True)
    except Exception as e:
        raise ValueError(f"Cannot read Excel file: {e}") from e
    return wb.sheetnames


def read_rows(
    content: bytes, filename: str, sheet_name: Optional[str] = None
) -> List[Dict[str, Any]]:
    if is_csv(filename):
        return read_csv(content)
    if is_excel(filename):
        return read_excel(content, sheet_name=sheet_name)
    if is_zip(filename):
        inner_content, inner_name = extract_first_supported_from_zip(content)
        return read_rows(inner_content, inner_name, sheet_name=sheet_name)
    raise ValueError(
        "Unsupported file format. Only .csv, .xlsx and .zip (containing .csv/.xlsx) are supported."
    )


def extract_first_supported_from_zip(content: bytes) -> tuple[bytes, str]:
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as zf:
            names = zf.namelist()
            supported = [
                n for n in names
                if (is_csv(n) or is_excel(n)) and not n.startswith("__")
            ]
            if not supported:
                raise ValueError(
                    "Zip does not contain a supported file (.csv or .xlsx)."
                )
            chosen = supported[0]
            with zf.open(chosen) as f:
                return f.read(), chosen
    except zipfile.BadZipFile as e:
        raise ValueError(f"Invalid zip file: {e}") from e


def validate_extension(filename: Optional[str], *, allow_zip: bool = True) -> None:
    if not filename:
        raise ValueError("File name is required")
    allowed = (".csv", ".xlsx")
    if allow_zip:
        allowed += (".zip",)
    if not filename.lower().endswith(allowed):
        raise ValueError(
            f"Unsupported file extension for {filename}. "
            f"Supported: {', '.join(allowed)}"
        )
