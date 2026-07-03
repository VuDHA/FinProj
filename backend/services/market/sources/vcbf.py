import datetime
import re
from io import BytesIO
from typing import Dict, List, Optional

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from common.models import Asset
from services.market.sources.base import Source
from services.market.sources.utils import parse_float, today


VCBF_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}


VCBF_FUNDS = {
    "VCBF-MGF": {
        "code": "mgf",
        "slug": "vcbf-midcap-growth-fund",
        "name": "VCBF Mid-Cap Growth Fund",
    },
    "VCBF-BCF": {
        "code": "bcf",
        "slug": "vcbf-blue-chip-fund",
        "name": "VCBF Blue Chip Fund",
    },
    "VCBF-AIF": {
        "code": "aif",
        "slug": "vcbf-active-income-fund",
        "name": "VCBF Active Income Fund",
    },
    "VCBF-TBF": {
        "code": "tbf",
        "slug": "vcbf-tactical-balanced-fund",
        "name": "VCBF Tactical Balanced Fund",
    },
    "VCBF-FIF": {
        "code": "fif",
        "slug": "vcbf-fixed-income-fund",
        "name": "VCBF Fixed Income Fund",
    },
}


class VcbfFundSource(Source):
    """NAV quỹ VCBF từ báo cáo Excel đăng trên website vcbf.com.

    VCBF đăng hàng ngày file Excel báo cáo NAV cho từng quỹ. Nguồn này phân tích
    các file Excel đó để lấy NAV mới nhất và lịch sử các ngày có sẵn trên website.
    """

    code = "vcbf"
    name = "VCBF"
    description = "NAV quỹ mở VCBF từ báo cáo Excel trên vcbf.com."
    supported_types = ["FUND"]
    supports_history = True
    supports_listing = True

    _LINKS_CACHE_TTL_SECONDS = 3600
    _links_cache: Dict[str, tuple[List[tuple], datetime.datetime]] = {}

    def _fund_info(self, symbol: str) -> Optional[dict]:
        return VCBF_FUNDS.get(symbol.upper())

    def _find_excel_links(self, symbol: str) -> List[tuple]:
        """Return list of (date, url) for available Excel NAV reports.

        Results are cached per symbol for _LINKS_CACHE_TTL_SECONDS to avoid
        re-scraping the same fund page on repeated price/history/detail calls.
        """
        cache_key = symbol.upper()
        now = datetime.datetime.now()
        cached = self._links_cache.get(cache_key)
        if cached:
            links, cached_at = cached
            if (now - cached_at).total_seconds() < self._LINKS_CACHE_TTL_SECONDS:
                return links

        fund = self._fund_info(symbol)
        if not fund:
            return []
        results = []
        try:
            url = f"https://www.vcbf.com/en/open-ended-funds/open-ended-funds-of-vcbf/{fund['slug']}/"
            r = requests.get(url, headers=VCBF_HEADERS, timeout=15)
            if r.status_code != 200:
                return []
            soup = BeautifulSoup(r.text, "html.parser")
            pattern = re.compile(rf"/images/\d{{4}}/vcb{fund['code']}_bc_(?:ngay|ky)_(\d{{8}})_1\.xlsx", re.IGNORECASE)
            for a in soup.find_all("a", href=True):
                match = pattern.search(a["href"])
                if match:
                    date_str = match.group(1)
                    try:
                        d = datetime.datetime.strptime(date_str, "%Y%m%d").date()
                    except Exception:
                        continue
                    full_url = a["href"]
                    if full_url.startswith("/"):
                        full_url = f"https://www.vcbf.com{full_url}"
                    results.append((d, full_url))
            results.sort(reverse=True)
            self._links_cache[cache_key] = (results, now)
        except Exception as e:
            print(f"[source vcbf] find links {symbol} error: {e}")
        return results

    def _parse_excel_nav(self, url: str) -> tuple[Optional[float], Optional[float]]:
        """Return (current_nav, previous_nav) per fund certificate from a VCBF Excel NAV report.

        VCBF uses different layouts for daily (bc_ngay) and periodic (bc_ky) reports,
        so we locate the row by the "per Fund Certificate" label and then read the first
        two non-empty numeric values that follow it.
        """
        try:
            r = requests.get(url, headers=VCBF_HEADERS, timeout=15)
            if r.status_code != 200:
                return (None, None)
            wb = load_workbook(BytesIO(r.content))
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                if not row:
                    continue
                text_idx = None
                for i, cell in enumerate(row):
                    if cell is not None and (
                        "per Fund Certificate" in str(cell) or "một chứng chỉ quỹ" in str(cell)
                    ):
                        text_idx = i
                        break
                if text_idx is None:
                    continue
                values = []
                for cell in row[text_idx + 1 :]:
                    if cell is None or str(cell).strip() == "":
                        continue
                    try:
                        val = float(str(cell).replace(",", ""))
                        if val > 0:
                            values.append(val)
                    except (ValueError, TypeError):
                        continue
                    if len(values) >= 2:
                        break
                if values:
                    return (values[0], values[1] if len(values) > 1 else None)
        except Exception as e:
            print(f"[source vcbf] parse excel {url} error: {e}")
        return (None, None)

    def fetch_price(self, asset: Asset) -> Optional[dict]:
        try:
            links = self._find_excel_links(asset.symbol)
            if not links:
                return None
            latest_date, latest_url = links[0]
            nav, prev_nav = self._parse_excel_nav(latest_url)
            if nav is None:
                return None
            change = (nav - prev_nav) if prev_nav is not None else 0.0
            change_percent = (change / prev_nav * 100) if prev_nav else 0.0
            return {
                "price": nav,
                "change": change,
                "change_percent": change_percent,
                "date": latest_date,
                "metadata": {"source": "vcbf", "report_url": latest_url},
            }
        except Exception as e:
            print(f"[source vcbf] price {asset.symbol} error: {e}")
        return None

    def fetch_history(
        self,
        symbol: str,
        asset_type: str,
        start: datetime.date,
        end: datetime.date,
    ) -> Dict[datetime.date, float]:
        result = {}
        try:
            links = self._find_excel_links(symbol)
            for d, url in links:
                if d < start or d > end:
                    continue
                nav, _ = self._parse_excel_nav(url)
                if nav is not None:
                    result[d] = nav
        except Exception as e:
            print(f"[source vcbf] history {symbol} error: {e}")
        return result

    def fetch_listing(self) -> List[dict]:
        results = []
        for symbol, info in VCBF_FUNDS.items():
            results.append(
                {
                    "symbol": symbol,
                    "name": info["name"],
                    "exchange": "VCBF",
                    "type": "FUND",
                    "fund_type": "FUND",
                    "metadata": {"slug": info["slug"]},
                }
            )
        return results

    def fetch_fund_detail(self, symbol: str) -> Optional[dict]:
        info = self._fund_info(symbol)
        if not info:
            return None
        nav = None
        links: List[tuple] = []
        try:
            links = self._find_excel_links(symbol)
            if links:
                nav, _ = self._parse_excel_nav(links[0][1])
        except Exception as e:
            print(f"[source vcbf] detail {symbol} error: {e}")
        return {
            "symbol": symbol,
            "name": info["name"],
            "fund_type": "FUND",
            "owner": "Vietcombank Fund Management Company Limited",
            "nav": nav or 0.0,
            "nav_update_at": links[0][0] if links else None,
        }
